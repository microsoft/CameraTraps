r"""
This program compiles an Excel spreadsheet for manually mapping dataset-specific
species names to a common taxonomy.

We are currently doing the counting of species here instead of as a part of the
Cosmos DB query
- see SDK issue in notes.

It first goes through the list of datasets in the `datasets` table to find out
which "species" are in each dataset and the count of its "occurrences" (each
sequence is counted as 1 if the class label is on the sequence; each image is
counted as 1 as well if the class label is on the image level; so a
sequence/image count mixture). This information is saved in a JSON file in the
`output_dir` for each dataset.

Once this information is collected, for each "species" in a dataset, it queries
the TOP 100 sequences where the "species" is in the list of class names at
either the sequence or the image level. It samples 7 of these TOP 500 sequences
(sequences returned by TOP may have little variety) and from each sequence
samples an image to surface as an example. The spreadsheet is then prepared,
adding a Bing search URL with the species class name as the query string and
fields to filter and fill in Excel.

Because querying for all species present in a dataset may take a long time, a
dataset is only queried if it does not yet have a JSON file in the `output_dir`.

Also, the spreadsheet creation step is only done for datasets in
DATASETS_TO_INCLUDE_IN_SPREADSHEET specified below. This is usually the datasets
just ingested that need their species names mapped next.

Leave out the flag `--query-species` if you only want to prepare the spreadsheet
using previously queried species presence result.


Example invocation:
    python taxonomy_mapping/species_by_dataset.py \
        --output-dir $HOME/megadb_query_results/species_by_dataset_trial \
        --query-species
"""

import argparse
from collections import Counter
from datetime import datetime
import json
import os
from random import sample
from typing import List, Optional
import urllib.parse

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from tqdm import tqdm

from data_management.megadb.megadb_utils import MegadbUtils


NUMBER_EXAMPLES_PER_SPECIES = 7
NUMBER_SEQUENCES_TO_QUERY = 500

DATASETS_TO_INCLUDE_IN_SPREADSHEET = [
    'ena24',
    'sulross_2018',
    'sulross_2019_spring',
    'sulross_kitfox',
    'idfg_swwlf_2019'
]


def query_species_by_dataset(megadb_utils: MegadbUtils,
                             output_dir: str) -> None:
    """For each dataset, creates a JSON file specifying species counts.

    Skips dataset if a JSON file for it already exists.
    """
    # which datasets are already processed?
    queried_datasets = set(
        i.split('.json')[0] for i in os.listdir(output_dir)
        if i.endswith('.json'))

    datasets_table = megadb_utils.get_datasets_table()
    dataset_names = [i for i in datasets_table if i not in queried_datasets]

    print(f'{len(queried_datasets)} datasets already queried. Querying species '
          f'in {len(dataset_names)} datasets...')

    for dataset_name in dataset_names:
        print(f'Querying dataset {dataset_name}...')

        # sequence-level query should be fairly fast, ~1 sec
        query_seq_level = '''
        SELECT VALUE seq.class
        FROM seq
        WHERE ARRAY_LENGTH(seq.class) > 0
            AND NOT ARRAY_CONTAINS(seq.class, "empty")
            AND NOT ARRAY_CONTAINS(seq.class, "__label_unavailable")
        '''
        results = megadb_utils.query_sequences_table(
            query_seq_level, partition_key=dataset_name)

        counter = Counter()
        for i in results:
            counter.update(i)

        # cases when the class field is on the image level (images in a sequence
        # that had different class labels, 'caltech' dataset is like this)
        # this query may take a long time, >1hr
        query_image_level = '''
        SELECT VALUE seq.images
        FROM sequences seq
        WHERE (
            SELECT VALUE COUNT(im)
            FROM im IN seq.images
            WHERE ARRAY_LENGTH(im.class) > 0
        ) > 0
        '''

        start = datetime.now()
        results_im = megadb_utils.query_sequences_table(
            query_image_level, partition_key=dataset_name)
        elapsed = (datetime.now() - start).seconds
        print(f'- image-level query took {elapsed}s')

        for seq_images in results_im:
            for im in seq_images:
                assert 'class' in im
                counter.update(im['class'])

        with open(os.path.join(output_dir, f'{dataset_name}.json'), 'w') as f:
            json.dump(counter, f, indent=2)


def get_example_images(megadb_utils: MegadbUtils, dataset_name: str,
                       class_name: str) -> List[Optional[str]]:
    """Gets SAS URLs for images of a particular class from a given dataset."""
    datasets_table = megadb_utils.get_datasets_table()

    # this query should be fairly fast, ~1 sec
    query_both_levels = f'''
    SELECT TOP {NUMBER_SEQUENCES_TO_QUERY} VALUE seq
    FROM seq
    WHERE ARRAY_CONTAINS(seq.class, "{class_name}")
        OR (SELECT VALUE COUNT(im)
            FROM im IN seq.images
            WHERE ARRAY_CONTAINS(im.class, "{class_name}")) > 0
    '''
    sequences = megadb_utils.query_sequences_table(
        query_both_levels, partition_key=dataset_name)

    num_samples = min(len(sequences), NUMBER_EXAMPLES_PER_SPECIES)
    sample_seqs = sample(sequences, num_samples)

    image_urls: List[Optional[str]] = []
    for seq in sample_seqs:
        sample_image = sample(seq['images'], 1)[0]  # sample 1 img per sequence
        img_path = MegadbUtils.get_full_path(
            datasets_table, dataset_name, sample_image['file'])
        img_path = urllib.parse.quote_plus(img_path)

        dataset_info = datasets_table[dataset_name]
        img_url = 'https://{}.blob.core.windows.net/{}/{}{}'.format(
            dataset_info["storage_account"],
            dataset_info["container"],
            img_path,
            dataset_info["container_sas_key"])
        image_urls.append(img_url)

    num_missing = NUMBER_EXAMPLES_PER_SPECIES - len(image_urls)
    if num_missing > 0:
        image_urls.extend([None] * num_missing)
    assert len(image_urls) == NUMBER_EXAMPLES_PER_SPECIES
    return image_urls


def make_spreadsheet(megadb_utils: MegadbUtils, output_dir: str) -> None:
    all_classes = set()
    class_in_multiple_ds = {}  # {class_name: bool}
    species_by_dataset = {}  # {dataset_name: {class_name: count}}

    classes_excluded = ['empty', 'car', 'vehicle', 'unidentified', 'unknown',
                        '__label_unavailable', 'error']

    # read species presence info from the JSON files for each dataset
    for file_name in os.listdir(output_dir):
        dataset_name, ext = os.path.splitext(file_name)
        if (ext != '.json') or (dataset_name not in DATASETS_TO_INCLUDE_IN_SPREADSHEET):
            continue
        print(f'Processing dataset {dataset_name}')

        with open(os.path.join(output_dir, file_name)) as f:
            class_counts = json.load(f)

        species_valid = {
            class_name: count for class_name, count in class_counts.items()
            if class_name not in classes_excluded
        }
        # has this class name appeared in a previous dataset?
        for class_name in species_valid:
            class_in_multiple_ds[class_name] = (class_name in all_classes)
            all_classes.add(class_name)

        species_by_dataset[dataset_name] = species_valid

    # columns to populate the spreadsheet
    col_order = [
        'dataset',
        'occurrences',  # count of sequences/images mixture with this class name
        'species_label',
        'bing_url',
        'is_common',  # is this class name seen already / need to be labeled again?
        'taxonomy_name',
        'common_name',
        'is_typo',  # there is a typo in the class name, but correct taxonomy name can be inferred
        'not_applicable',  # labels like "human-cattle" where a taxonomy name would not be applicable
        'other_notes',  # other info in the class name, like male/female
        'is_new'  # not in pervious versions of this spreadsheet
    ]
    for i in range(NUMBER_EXAMPLES_PER_SPECIES):
        col_order.append(f'example{i + 1}')
    col_order.append('example_mislabeled')

    rows = []
    bing_prefix = 'https://www.bing.com/search?q='
    for dataset_name, species_count in species_by_dataset.items():
        print(dataset_name)

        # sort by descending species count
        species_count_tups = sorted(species_count.items(),
                                    key=lambda x: x[1], reverse=True)
        for class_name, class_count in tqdm(species_count_tups):
            row = dict(
                dataset=dataset_name,
                occurrences=class_count,
                species_label=class_name,
                bing_url=bing_prefix + urllib.parse.quote_plus(class_name),
                is_common=class_in_multiple_ds[class_name],
                taxonomy_name='',
                common_name='',
                is_typo='',
                other_notes='',
                not_applicable='',
                is_new=True,
                example_mislabeled='')

            example_images_sas_urls = get_example_images(
                megadb_utils, dataset_name, class_name)
            for i, url in enumerate(example_images_sas_urls):
                row[f'example{i + 1}'] = url

            rows.append(row)

    # make the spreadsheet
    spreadsheet = pd.DataFrame(data=rows, columns=col_order)
    print(spreadsheet.head(5))
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(spreadsheet, index=False, header=True):
        ws.append(r)

    # hyperlink Bing search URLs
    for i_row, cell in enumerate(ws['D']):  # TODO hardcoded column number
        if i_row > 0:
            cell.hyperlink = cell.value
            cell.style = 'Hyperlink'

    # hyperlink example image SAS URLs
    # TODO hardcoded columns: change if # of examples or col_order changes
    sas_cols = [ws['L'], ws['M'], ws['N'], ws['O'], ws['P'], ws['Q'], ws['R']]
    assert len(sas_cols) == NUMBER_EXAMPLES_PER_SPECIES

    for i_example, ws_col in enumerate(sas_cols):
        for i_row, cell in enumerate(ws_col):
            if i_row > 0 and cell.value is not None:
                if not isinstance(cell.value, str):
                    print(f'WARNING cell.value is {cell.value}, '
                          f'type is {type(cell.value)}')
                    continue
                cell.hyperlink = cell.value
                cell.value = f'example{i_example + 1}'
                cell.style = 'Hyperlink'

    date = datetime.now().strftime('%Y_%m_%d')
    wb.save(os.path.join(output_dir, f'species_by_dataset_{date}.xlsx'))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-o', '--output-dir', required=True,
        help='Path to directory where the JSONs containing species count for '
             'each dataset live')
    parser.add_argument(
        '-q', '--query-species', action='store_true',
        help='Query what species are present in a dataset. '
             'Otherwise, create a spreadsheet for labeling the taxonomy.')
    args = parser.parse_args()

    assert 'COSMOS_ENDPOINT' in os.environ and 'COSMOS_KEY' in os.environ

    os.makedirs(args.output_dir, exist_ok=True)

    megadb_utils = MegadbUtils()

    if args.query_species:
        query_species_by_dataset(megadb_utils, args.output_dir)
    make_spreadsheet(megadb_utils, args.output_dir)


if __name__ == '__main__':
    main()
